import gurobipy as gp
from gurobipy import GRB
import io
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import asyncio
import threading
import tempfile
import os
from openpyxl.styles import Font

class CaptureOutput(list):
    """Class to capture stdout and stderr."""
    def __enter__(self):
        self._stdout = sys.stdout
        self._stderr = self._stderr = io.StringIO()
        sys.stdout = self._stderr
        sys.stderr = self._stderr
        return self

    def __exit__(self, *args):
        self.extend(self._stderr.getvalue().splitlines())
        sys.stdout = self._stdout
        sys.stderr = self._stderr

def execute_model_from_file(model_file: str, time_limit: float, temp_dir: str) -> str:
    # 從文件讀取模型
    model = gp.read(model_file)

    # 設置最大執行時間
    if time_limit > 0:
        model.setParam(GRB.Param.TimeLimit, time_limit)

    # 捕獲 Gurobi 的日誌輸出
    with CaptureOutput() as output:
        # 求解模型
        model.optimize()
        model.printStats()

    log_output = output._stderr.getvalue().splitlines()

    # 顯示捕獲的日誌輸出到控制台
    print('')
    print("\n".join(["-"*80, "Gurobi Output Log", "-"*80]))
    for line in log_output:
        print(line)
    print("-"*80 + "\n")

    # 將結果寫入臨時文件
    temp_file = os.path.join(temp_dir, f"temp_{os.path.basename(model_file)}.txt")
    with open(temp_file, "w") as f:
        for v in model.getVars():
            f.write(f"{v.VarName}: {v.X}\n")
        f.write("\n" + "-"*80 + "\n")  # 日誌輸出開始的標記
        f.write("\n".join(log_output))
        # 添加附加信息
        f.write(f"\nObjective Value: {model.ObjVal}\n")
        f.write(f"Solve Time (seconds): {model.Runtime}\n")
        f.write(f"Iterations: {model.IterCount}\n")
        f.write(f"Node Count: {model.NodeCount}\n")
    
    return temp_file

def read_temp_files(temp_dir: str) -> dict:
    results = {}
    for temp_file in os.listdir(temp_dir):
        if temp_file.startswith("temp_") and temp_file.endswith(".txt"):
            model_name = temp_file.replace('temp_', '').replace('.txt', '')
            data = {
                'Variable': [],
                'Value': [],
                'Log': [],
            }
            with open(os.path.join(temp_dir, temp_file), 'r') as f:
                lines = f.readlines()
                is_log = False
                for line in lines:
                    if is_log:
                        data['Log'].append(line.strip())
                    elif line.strip() == "-"*80:
                        is_log = True
                    elif ': ' in line:
                        var_name, value = line.strip().split(': ')
                        data['Variable'].append(var_name)
                        data['Value'].append(float(value))
            results[model_name] = data
            os.remove(os.path.join(temp_dir, temp_file))
    
    # 確保刪除空的暫存資料夾
    if not os.listdir(temp_dir):
        os.rmdir(temp_dir)
    
    return results

def save_to_excel(results: dict, filename: str) -> None:
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        model_info_data = {'Model Number': [], 'Model File': []}
        
        for idx, (model_name, data) in enumerate(results.items(), start=1):
            # 準備變數和值數據
            df = pd.DataFrame({
                'Variable': data['Variable'],
                'Value': data['Value']
            })

            # 捕獲的日誌輸出，分段保存
            df_log = pd.DataFrame({'Log': data['Log']})

            model_info_data['Model Number'].append(idx)
            model_info_data['Model File'].append(model_name)

            # 寫入數據到不同的工作表
            df.to_excel(writer, sheet_name=f'Model_{idx}_Solution', index=False, startrow=1)
            df_log.to_excel(writer, sheet_name=f'Model_{idx}_Log', index=False, startrow=1)

            # 在工作表的第一行添加模型文件名
            sheet = writer.sheets[f'Model_{idx}_Solution']
            sheet.cell(row=1, column=1, value=f'Model File: {model_name}')

            sheet = writer.sheets[f'Model_{idx}_Log']
            sheet.cell(row=1, column=1, value=f'Model File: {model_name}')

            # 设置 "Time limit reached" 为红色粗体
            red_bold_font = Font(color="FF0000", bold=True)
            for row_idx, log_entry in enumerate(data['Log'], start=2):
                if "Time limit reached" in log_entry:
                    cell = sheet.cell(row=row_idx + 1, column=1)  # +1 因为日志输出从第2行开始
                    cell.font = red_bold_font

        # 寫入模型編號和文件名對應表
        df_model_info = pd.DataFrame(model_info_data)
        df_model_info.to_excel(writer, sheet_name='Model_Index', index=False)

def main():
    # 建立主窗口
    root = tk.Tk()
    root.title("Gurobi Model Runner")
    # 禁用視窗拉伸
    root.resizable(False, False)

    file_entries = []
    file_entry_widgets = []
    browse_buttons = []
    solve_thread = None
    results = {}
    temp_dir = None

    def add_file_entry():
        frame = tk.Frame(root, padx=10, pady=5)
        frame.grid(row=len(file_entries) + 6, column=0, sticky=tk.W, columnspan=5)
        
        model_files = tk.StringVar()
        file_entries.append(model_files)

        label = tk.Label(frame, text="模型檔案路徑:")
        label.grid(row=0, column=0, sticky=tk.W)

        entry = tk.Entry(frame, textvariable=model_files, width=50)
        entry.grid(row=0, column=1, padx=5, pady=5)
        file_entry_widgets.append(entry)

        button_browse = tk.Button(frame, text="瀏覽...", command=lambda: select_files(model_files))
        button_browse.grid(row=0, column=2, padx=5, pady=5)
        browse_buttons.append(button_browse)

    # 文件選擇函數
    def select_files(model_files):
        file_paths = filedialog.askopenfilenames(filetypes=[("Model Files", "*.lp *.mps *.json")])
        if file_paths:
            model_files.set(";".join(file_paths))

    # 執行模型求解
    async def solve_models():
        try:
            nonlocal temp_dir
            temp_dir = tempfile.mkdtemp()

            model_file_paths = [path for entry in file_entries for path in entry.get().split(";") if entry.get()]
            if not model_file_paths:
                messagebox.showwarning("Warning", "請至少選擇一個模型檔案")
                return

            toggle_buttons(state=tk.DISABLED)

            progress['value'] = 0
            progress['maximum'] = len(model_file_paths)
            progress_label.config(text=f"0/{len(model_file_paths)} (0.00%)")
            
            time_limit = float(time_limit_var.get())

            for i, file_path in enumerate(model_file_paths):
                model_name = file_path.split("/")[-1]
                current_model.set(f"正在處理: {model_name}")
                loop = asyncio.get_event_loop()
                temp_file = await loop.run_in_executor(None, execute_model_from_file, file_path, time_limit, temp_dir)
                
                if temp_file:
                    if "Time limit reached" in open(temp_file).read():
                        result = f"{model_name}: 達到執行時限"
                    else:
                        result = f"{model_name}: 模型已解決"
                else:
                    result = f"{model_name}: 模型求解失敗"
                
                progress['value'] = i + 1

                # 更新進度信息
                progress_label.config(text=f"{i + 1}/{len(model_file_paths)} ({(i + 1) / len(model_file_paths) * 100:.2f}%)")
                result_text.set(result)
                await asyncio.sleep(0)  # 让出控制权以保持GUI响应

            current_model.set("所有模型皆已處理完畢，請記得存檔！")
            results['data'] = read_temp_files(temp_dir)
            toggle_buttons(state=tk.NORMAL)
            save_button.config(state=tk.NORMAL)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            toggle_buttons(state=tk.NORMAL)
            save_button.config(state=tk.DISABLED)
        
        # 清空路徑框
        for entry in file_entries:
            entry.set("")

    # 儲存結果為 Excel
    def save_result_to_excel():
        try:
            if not results.get('data'):
                messagebox.showwarning("Warning", "沒有可保存的模型結果")
                return

            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if file_path:
                save_to_excel(results['data'], file_path)
                messagebox.showinfo("Info", "結果已成功保存")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # 切換按鈕和輸入框狀態
    def toggle_buttons(state=tk.NORMAL):
        for widget in [button_add_file, button_solve, save_button]:
            widget.config(state=state)
        for entry in file_entry_widgets:
            entry.config(state=state)
        for button in browse_buttons:
            button.config(state=state)
        time_limit_entry.config(state=state)

    # 結果變數
    result_text = tk.StringVar()
    current_model = tk.StringVar()
    time_limit_var = tk.StringVar(value="0")

    # 建立UI元件
    current_model_label = tk.Label(root, textvariable=current_model, justify=tk.LEFT)
    current_model_label.grid(row=0, column=0, columnspan=5)

    progress_frame = tk.Frame(root, padx=10, pady=5)
    progress_frame.grid(row=1, column=0, columnspan=5)

    progress = ttk.Progressbar(progress_frame, orient='horizontal', length=350, mode='determinate')
    progress.grid(row=0, column=0, columnspan=4, sticky=tk.E)

    progress_label = tk.Label(progress_frame, text="0/0 (0.00%)", justify=tk.CENTER)
    progress_label.grid(row=0, column=4, sticky=tk.W)

    result_label = tk.Label(root, textvariable=result_text, justify=tk.LEFT)
    result_label.grid(row=2, column=0, columnspan=5)

    time_limit_frame = tk.Frame(root, padx=10, pady=5)
    time_limit_frame.grid(row=3, column=2, columnspan=1, sticky=tk.W)

    time_limit_label = tk.Label(time_limit_frame, text="單一模型最大執行時間 (秒，0表示不限制):")
    time_limit_label.grid(row=0, column=0, sticky=tk.E, columnspan=3)
    
    time_limit_entry = tk.Entry(time_limit_frame, textvariable=time_limit_var, width=10)
    time_limit_entry.grid(row=0, column=3, sticky=tk.W)

    button_frame = tk.Frame(root, padx=10, pady=5)
    button_frame.grid(row=4, column=0, columnspan=5, sticky=tk.W)

    button_add_file = tk.Button(button_frame, text="添加模型路徑輸入框", command=add_file_entry)
    button_add_file.grid(row=0, column=0, padx=5, pady=5)

    button_solve = tk.Button(button_frame, text="求解模型", command=lambda: start_solve_thread())
    button_solve.grid(row=0, column=1, padx=5, pady=5)

    save_button = tk.Button(button_frame, text="將結果輸出為 Excel 檔（.xlsx）", command=save_result_to_excel, state=tk.DISABLED)
    save_button.grid(row=0, column=2, padx=5, pady=5)

    add_file_entry()  # 添加第一個文件輸入框

    def start_solve_thread():
        global solve_thread
        solve_thread = threading.Thread(target=asyncio.run, args=(solve_models(),))
        solve_thread.start()

    root.mainloop()

if __name__ == "__main__":
    main()
